#!/usr/bin/env python3
"""Validate and export the complete football workbook for GitHub Pages."""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


SHEET_CONFIG = {
    "数据看板": {"slug": "dashboard", "headerRow": None, "dataStartRow": 1, "mode": "grid"},
    "比赛台账": {"slug": "matches", "headerRow": 4, "dataStartRow": 5, "mode": "table"},
    "赛事字典": {"slug": "dictionary", "headerRow": 4, "dataStartRow": 5, "mode": "table"},
    "维度战绩": {"slug": "dimensions", "headerRow": None, "dataStartRow": 1, "mode": "grid"},
}


def json_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float):
        if value == int(value):
            return int(value)
        return round(value, 10)
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def row_has_value(values):
    return any(value not in (None, "") for value in values)


def trim_trailing_empty(values):
    values = list(values)
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def normalized_excel_datetime(value, epoch):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch)
        except (OverflowError, ValueError):
            return None
        if isinstance(converted, datetime):
            return converted
        if isinstance(converted, date):
            return datetime.combine(converted, time.min)
    return None


def validate_ledger(values_sheet, formula_sheet, epoch):
    records = []
    blank_seen = False
    gap_rows = []
    formula_rows = []

    for row_number in range(5, values_sheet.max_row + 1):
        values = [values_sheet.cell(row_number, column).value for column in range(1, 19)]
        nonempty = row_has_value(values)
        if not nonempty:
            blank_seen = True
            continue
        if blank_seen:
            gap_rows.append(row_number)

        kickoff = normalized_excel_datetime(values[1], epoch)
        if kickoff is None:
            raise ValueError(f"比赛台账 B{row_number} 不是有效开球时间")

        source_formula = formula_sheet.cell(row_number, 1).value
        profit_formula = formula_sheet.cell(row_number, 16).value
        if not (isinstance(source_formula, str) and source_formula.startswith("=")):
            formula_rows.append(f"A{row_number}")
        if not (isinstance(profit_formula, str) and profit_formula.startswith("=")):
            formula_rows.append(f"P{row_number}")
        records.append((row_number, kickoff))

    if gap_rows:
        raise ValueError(f"比赛台账存在空白夹层，首个异常行：{gap_rows[0]}")
    if formula_rows:
        raise ValueError(f"比赛台账公式缺失，首个异常单元格：{formula_rows[0]}")

    for previous, current in zip(records, records[1:]):
        if previous[1] < current[1]:
            raise ValueError(
                f"比赛台账未按开球时间降序：B{previous[0]}={previous[1]} < B{current[0]}={current[1]}"
            )

    return {
        "records": len(records),
        "firstKickoff": records[0][1].strftime("%Y-%m-%d %H:%M") if records else None,
        "lastKickoff": records[-1][1].strftime("%Y-%m-%d %H:%M") if records else None,
        "descending": True,
        "formulaColumns": ["A", "P"],
    }


def export_sheet(values_sheet, formula_sheet, config, output_path, epoch):
    rows = []
    last_nonempty_row = 0
    for row_number in range(1, values_sheet.max_row + 1):
        data_values = [values_sheet.cell(row_number, column).value for column in range(1, values_sheet.max_column + 1)]
        if not row_has_value(data_values):
            continue
        cells = []
        for column in range(1, values_sheet.max_column + 1):
            value = data_values[column - 1]
            formula = formula_sheet.cell(row_number, column).value
            if value is None and isinstance(formula, str) and formula.startswith("="):
                value = formula
            if values_sheet.title == "比赛台账" and row_number >= 5 and column in (2, 9):
                value = normalized_excel_datetime(value, epoch) or value
            cells.append(json_value(value))
        cells = trim_trailing_empty(cells)
        if cells:
            rows.append({"number": row_number, "cells": cells})
            last_nonempty_row = row_number

    payload = {
        "name": values_sheet.title,
        "slug": config["slug"],
        "mode": config["mode"],
        "usedRange": f"A1:{get_column_letter(values_sheet.max_column)}{last_nonempty_row}",
        "rowCount": last_nonempty_row,
        "columnCount": values_sheet.max_column,
        "headerRow": config["headerRow"],
        "dataStartRow": config["dataStartRow"],
        "columns": [get_column_letter(index) for index in range(1, values_sheet.max_column + 1)],
        "rows": rows,
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return {
        "name": values_sheet.title,
        "slug": config["slug"],
        "file": f"./workbook/{output_path.name}",
        "mode": config["mode"],
        "usedRange": payload["usedRange"],
        "rowCount": last_nonempty_row,
        "columnCount": values_sheet.max_column,
        "headerRow": config["headerRow"],
        "dataStartRow": config["dataStartRow"],
        "nonemptyRows": len(rows),
        "sizeBytes": output_path.stat().st_size,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--download-dir", type=Path)
    args = parser.parse_args()

    probe = load_workbook(args.workbook, read_only=True)
    missing = [name for name in SHEET_CONFIG if name not in probe.sheetnames]
    probe.close()
    if missing:
        raise ValueError(f"工作簿缺少工作表：{', '.join(missing)}")

    values_workbook = load_workbook(args.workbook, read_only=False, data_only=True)
    formula_workbook = load_workbook(args.workbook, read_only=False, data_only=False)
    validation = validate_ledger(
        values_workbook["比赛台账"], formula_workbook["比赛台账"], values_workbook.epoch
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)
    sheets = []
    for sheet_name, config in SHEET_CONFIG.items():
        output_path = args.output_dir / f"{config['slug']}.json"
        sheets.append(
            export_sheet(
                values_workbook[sheet_name],
                formula_workbook[sheet_name],
                config,
                output_path,
                values_workbook.epoch,
            )
        )
    values_workbook.close()
    formula_workbook.close()

    workbook_bytes = args.workbook.read_bytes()
    download_path = "./files/足球比赛自动分析台账.xlsx"
    if args.download_dir:
        args.download_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.workbook, args.download_dir / "足球比赛自动分析台账.xlsx")

    manifest = {
        "workbook": "足球比赛自动分析台账.xlsx",
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sha256": hashlib.sha256(workbook_bytes).hexdigest(),
        "sizeBytes": len(workbook_bytes),
        "downloadPath": download_path,
        "validation": validation,
        "sheets": sheets,
    }
    manifest_path = args.output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print(json.dumps(manifest, ensure_ascii=False))


if __name__ == "__main__":
    main()
