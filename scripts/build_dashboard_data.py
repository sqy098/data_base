#!/usr/bin/env python3
"""Build the public dashboard data bundle from the football ledger workbook."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from numbers import Number
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


HEADERS = [
    "sourceKey",
    "kickoff",
    "league",
    "home",
    "away",
    "handicapPick",
    "totalPick",
    "optional",
    "frozenAt",
    "status",
    "score",
    "handicapResult",
    "totalResult",
    "handicapProfit",
    "totalProfit",
    "profit",
    "source",
    "note",
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def numeric(value):
    if isinstance(value, Number) and not isinstance(value, bool):
        return float(value)
    return None


def fmt_datetime(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value).strip() if value else ""


def normalized_excel_datetime(value, epoch):
    if isinstance(value, datetime):
        return value
    if isinstance(value, Number) and not isinstance(value, bool):
        try:
            converted = from_excel(value, epoch)
        except (OverflowError, ValueError):
            return value
        return converted if isinstance(converted, datetime) else value
    return value


def add_months(year, month, delta):
    month_index = year * 12 + (month - 1) + delta
    return month_index // 12, month_index % 12 + 1


def summarize_market(rows, value_key):
    values = [numeric(row[value_key]) for row in rows]
    values = [value for value in values if value is not None]
    return {
        "bets": len(values),
        "profit": round(sum(values), 4),
        "positiveRate": (sum(value > 0 for value in values) / len(values)) if values else 0,
    }


def ranking(rows, value_key, dimension):
    buckets = defaultdict(list)
    for row in rows:
        value = numeric(row[value_key])
        if value is None:
            continue
        names = [row[dimension]] if dimension == "league" else [row["home"], row["away"]]
        for name in names:
            if name:
                buckets[str(name)].append(value)

    results = []
    for name, values in buckets.items():
        profit = round(sum(values), 4)
        bets = len(values)
        results.append(
            {
                "name": name,
                "matches": bets,
                "bets": bets,
                "profit": profit,
                "positiveRate": sum(value > 0 for value in values) / bets,
                "average": round(profit / bets, 2),
            }
        )
    return sorted(
        results,
        key=lambda item: (
            -item["profit"],
            -item["positiveRate"],
            -item["bets"],
            item["name"],
        ),
    )


def match_payload(row):
    handicap_profit = numeric(row["handicapProfit"])
    total_profit = numeric(row["totalProfit"])
    combined = None
    if handicap_profit is not None or total_profit is not None:
        combined = round((handicap_profit or 0) + (total_profit or 0), 4)
    source_key = str(row["sourceKey"]) if row["sourceKey"] else "|".join(
        [fmt_datetime(row["kickoff"])[:10], str(row["home"]), str(row["away"])]
    )
    return {
        "sourceKey": source_key,
        "date": fmt_datetime(row["kickoff"]),
        "frozenAt": fmt_datetime(row["frozenAt"]),
        "league": str(row["league"]),
        "home": str(row["home"]),
        "away": str(row["away"]),
        "handicapPick": str(row["handicapPick"]),
        "totalPick": str(row["totalPick"]),
        "optional": str(row["optional"]),
        "score": str(row["score"]),
        "status": str(row["status"]),
        "handicapResult": str(row["handicapResult"]),
        "totalResult": str(row["totalResult"]),
        "handicapProfit": handicap_profit,
        "totalProfit": total_profit,
        "profit": combined,
        "source": str(row["source"]),
        "note": str(row["note"]),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook["比赛台账"]
    rows = []
    for values in sheet.iter_rows(min_row=5, max_col=len(HEADERS), values_only=True):
        if not any(value is not None for value in values):
            continue
        row = {key: clean(value) for key, value in zip(HEADERS, values)}
        row["kickoff"] = normalized_excel_datetime(row["kickoff"], workbook.epoch)
        row["frozenAt"] = normalized_excel_datetime(row["frozenAt"], workbook.epoch)
        if not row["kickoff"] or not row["home"] or not row["away"]:
            continue
        rows.append(row)

    rows.sort(
        key=lambda row: row["kickoff"] if isinstance(row["kickoff"], datetime) else datetime.min,
        reverse=True,
    )

    handicap = summarize_market(rows, "handicapProfit")
    over_under = summarize_market(rows, "totalProfit")
    settled = sum(str(row["status"]) == "已结算" for row in rows)

    latest_kickoff = max(
        (row["kickoff"] for row in rows if isinstance(row["kickoff"], datetime)),
        default=datetime.now(),
    )
    monthly = []
    for offset in range(-5, 1):
        year, month = add_months(latest_kickoff.year, latest_kickoff.month, offset)
        month_rows = [
            row
            for row in rows
            if isinstance(row["kickoff"], datetime)
            and row["kickoff"].year == year
            and row["kickoff"].month == month
        ]
        monthly.append(
            {
                "label": f"{month}月",
                "handicapProfit": round(
                    sum(value for row in month_rows if (value := numeric(row["handicapProfit"])) is not None),
                    4,
                ),
                "totalProfit": round(
                    sum(value for row in month_rows if (value := numeric(row["totalProfit"])) is not None),
                    4,
                ),
            }
        )

    latest_frozen = max(
        (row["frozenAt"] for row in rows if isinstance(row["frozenAt"], datetime)),
        default=None,
    )

    pending_matches = [match_payload(row) for row in rows if str(row["status"]) != "已结算"]
    settled_matches = [match_payload(row) for row in rows if str(row["status"]) == "已结算"]

    update_text = latest_frozen.strftime("%Y-%m-%d %H:%M") if latest_frozen else "待更新"
    data = {
        "updatedAt": f"数据更新：{update_text}（北京时间） · 最新台账 {len(rows):,} 场",
        "totals": {
            "matches": len(rows),
            "settled": settled,
            "combinedProfit": round(handicap["profit"] + over_under["profit"], 4),
            "handicap": handicap,
            "overUnder": over_under,
        },
        "monthly": monthly,
        "leagueHandicap": ranking(rows, "handicapProfit", "league"),
        "leagueTotals": ranking(rows, "totalProfit", "league"),
        "teamHandicap": ranking(rows, "handicapProfit", "team"),
        "teamTotals": ranking(rows, "totalProfit", "team"),
        "pendingMatches": pending_matches,
        "settledMatches": settled_matches,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    args.output.write_text(
        "/* Generated from 足球比赛自动分析台账.xlsx. */\nwindow.DASHBOARD_DATA=" + payload + ";\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "matches": len(rows),
                "settled": settled,
                "handicap": handicap,
                "overUnder": over_under,
                "pendingMatches": len(pending_matches),
                "settledMatches": len(settled_matches),
                "output": str(args.output),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
